import openpyxl


def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_row


def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_column


def readdata(file, sheet, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.cell(row=rowno, column=colno).value


def writedata(file, sheet, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    sheet.cell(row=rowno, column=colno).value = data
    workbook.save()

